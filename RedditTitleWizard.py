import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
import threading
import pandas as pd
import random
from difflib import SequenceMatcher
import praw
import openpyxl
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.naive_bayes import MultinomialNB
# data
PATH = r"Data.xlsx"
df = pd.read_excel(PATH)

# reddit
reddit = praw.Reddit(
    client_id="***",
    client_secret="***",
    user_agent="***",
)

# loading bar
def start_loading():
    progress_bar["mode"] = "determinate"
    progress_bar["value"] = 0
    root.update_idletasks()

def stop_loading():
    progress_bar["mode"] = "indeterminate"
    progress_bar.stop()
    

def collect_data_to_excel(PATH):
    subreddit_name = "****"
    current_row = 2
    subreddit = reddit.subreddit(subreddit_name)
    num_posts = subreddit.subscribers
    workbook = openpyxl.load_workbook(PATH)
    worksheet = workbook['Sheet1']
    # Get existing titles from the Excel file
    existing_titles = set(worksheet.cell(row=row, column=1).value for row in range(2, worksheet.max_row + 1))

    for submission in reddit.subreddit(subreddit_name).new(limit=None):
        if submission.title not in existing_titles:  # Check if the title already exists
            worksheet = workbook['Sheet1']
            worksheet.cell(row=current_row, column=1, value=submission.title)
            worksheet.cell(row=current_row, column=2, value=submission.score)
            worksheet.cell(row=current_row, column=3, value=submission.num_comments)
            worksheet.cell(row=current_row, column=4, value=len(submission.all_awardings))
            worksheet.cell(row=current_row, column=5, value=submission.url)
            worksheet.cell(row=current_row, column=6, value=submission.link_flair_text)
            workbook.save(PATH)
            current_row += 1
    workbook.close()

def predict_data(input_string):
    global df, PATH
    collect_data_to_excel(PATH)
    X = df['Title']
    features = ["Score", "Comments", "Awards"]
    information = []
    
    for x in range(3):
        y = df[features[x]]
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.5, random_state=50)
        vectorizer = CountVectorizer()
        X_train_vec = vectorizer.fit_transform(X_train)
        X_test_vec = vectorizer.transform(X_test)
        model = MultinomialNB()
        model.fit(X_train_vec, y_train)
        input_vec = vectorizer.transform([input_string.strip()])
        prediction = model.predict(input_vec)
        information.append(prediction[0])
    return information

# input
def get_input():
    input_text = entryName.get()
    suggested_title = makeTitle(input_text) 
    entryNameSuggest.delete(0, tk.END)
    entryNameSuggest.insert(0, suggested_title)
    labelInput.config(text="point: {} comments: {} Awards: {}".format(predict_data(input_text)[0],predict_data(input_text)[1],predict_data(input_text)[2]))
    labeSugest.config(text="SUGESTED point: {} comments: {} Awards: {}".format(predict_data(suggested_title)[0],predict_data(suggested_title)[1],predict_data(suggested_title)[2]))
    start_loading()
    prediction_thread = threading.Thread(target=predict_and_update_labels, args=(input_text,))
    prediction_thread.start()

def predict_and_update_labels(input_text):
    information = predict_data(input_text)
    labelInput.config(text="point: {} comments: {} Awards: {}".format(information[0], information[1], information[2]))
    suggested_information = predict_data(entryNameSuggest.get())
    labeSugest.config(text="SUGGESTED point: {} comments: {} Awards: {}".format(
        suggested_information[0], suggested_information[1], suggested_information[2]))

# images
def open_file_dialog():
    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp")])
    if file_path:
        print("Selected file:", file_path)

# create/find
def makeTitle(input):
    global df
    similarity_threshold = 0.50
    title_scores = []
    input_words = input.split()
    for index, row in df.iterrows():
        title = row['Title']
        score = row['Score']
        comments = row['Comments']
        awards = row['Awards']
        title_words = title.split()
        similarity_scores = [SequenceMatcher(None, input_word, title_word).ratio() for input_word in input_words for title_word in title_words]
        average_similarity_score = sum(similarity_scores) / len(similarity_scores)
        if average_similarity_score >= similarity_threshold:
            title_scores.append([title, score, comments, awards])

    if title_scores:
        max_title_score = max(title_scores, key=lambda x: (x[1], x[2], x[3]))
        return max_title_score[0]
    else:
        random_row = df.sample(n=1)
        random_title = random_row['Title'].values[0]
        return random_title

# tinker window
root = tk.Tk()
# main
root.title("r/randomscreenshot")
root.geometry("600x600")
root.config(bg="black")
# labels
labelName = tk.Label(root, text="Enter post title:")
labelName.grid(row=0, column=0)
labelNameSugest = tk.Label(root, text="would you like to try:")
labelNameSugest.grid(row=1, column=0)
labelInput = tk.Label(root, text="point: {} comments: {} Awards: {}".format(0,0,0))
labelInput.grid(row=4, column=1)
labeSugest = tk.Label(root, text="SUGESTED point: {} comments: {} Awards: {}".format(0,0,0))
labeSugest.grid(row=5, column=1)
# Entry widgets
entryName = tk.Entry(root,)
entryName.grid(row=0, column=2)
entryNameSuggest = tk.Entry(root)
entryNameSuggest.grid(row=1, column=2)
upload_button = tk.Button(root, text="Upload Photo", command=open_file_dialog)
upload_button.grid(row=3,column = 1)
# buttons
button = tk.Button(root, text="Submit", command=get_input)
button.grid(row=6, columnspan=1)
progress_bar = ttk.Progressbar(root, orient="horizontal", length=200, mode="indeterminate")
progress_bar.grid(row=7, columnspan=1)
root.mainloop()
