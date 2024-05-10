import openpyxl
import pandas as pd
import praw
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.naive_bayes import MultinomialNB
import nltk
from sklearn.feature_extraction.text import TfidfVectorizer
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from sklearn.metrics.pairwise import cosine_similarity

nltk.download('stopwords')
nltk.download('punkt')

path = r"C:\Users\danie\OneDrive\Desktop\Data.xlsx"
workbook = openpyxl.load_workbook(path)

reddit = praw.Reddit(
    client_id="***",
    client_secret="***",
    user_agent="***",
)
worksheet = workbook['Sheet1']

# --- add data
current_row = 2
subreddit_name = "randomscreenshot"
subreddit = reddit.subreddit(subreddit_name)
num_posts = subreddit.subscribers
print(f"Looking through {num_posts} posts....")

# Get existing titles from the Excel file
existing_titles = set(worksheet.cell(row=row, column=1).value for row in range(2, worksheet.max_row + 1))

for submission in reddit.subreddit(subreddit_name).new(limit=None):
    if submission.title not in existing_titles:  # Check if the title already exists
        worksheet = workbook['Sheet1']
        worksheet.cell(row=current_row, column=1, value=submission.title)
        worksheet.cell(row=current_row, column=2, value=submission.score)
        worksheet.cell(row=current_row, column=3, value=submission.num_comments)
        worksheet.cell(row=current_row, column=4, value=len(submission.all_awardings))
        worksheet.cell(row=current_row, column=5, value=submission.url)
        worksheet.cell(row=current_row, column=6, value=submission.link_flair_text)
        workbook.save(path)
        current_row += 1

workbook.close()

input_string = input("Enter a Title > ")
data = pd.read_excel(path)

# --- perdict data
X = data['Title']
features  = ["Score","Comments","Awards"]
information = []
for x in range(3):
    y = data[features[x]]
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.5, random_state=50)
    vectorizer = CountVectorizer()
    X_train_vec = vectorizer.fit_transform(X_train)
    X_test_vec = vectorizer.transform(X_test)
    model = MultinomialNB()
    model.fit(X_train_vec, y_train)
    input_vec = vectorizer.transform([input_string.strip()])
    prediction = model.predict(input_vec)
    information.append(prediction[0])

print(f"perdicted title Data with the current title input: {input_string} \nscore: {information[0]} \nComments {information[1]} \nAwards {information[2]}")

# suggest a title
stop_words = set(stopwords.words("english"))
def preprocess_text(text):
    tokens = word_tokenize(text.lower())
    filtered_tokens = [token for token in tokens if token.isalnum() and token not in stop_words]
    return " ".join(filtered_tokens)

data["Title"] = data["Title"].apply(preprocess_text)
input_title = preprocess_text(input_string)
vectorizer = TfidfVectorizer()
tfidf_matrix = vectorizer.fit_transform(data["Title"])
input_vector = vectorizer.transform([input_title])
cosine_similarities = cosine_similarity(input_vector, tfidf_matrix)
data["Combined_Score"] = cosine_similarities[0] + data["Score"]
best_suggestion = data.loc[data["Combined_Score"].idxmax()]["Title"]
max_score = data["Combined_Score"].max()
print(f"Input Title: {input_title}")
print(f"Best Suggestion: {best_suggestion}")
print(f"Max Combined Score: {max_score}")
